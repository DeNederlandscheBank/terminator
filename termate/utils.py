# -*- coding: utf-8 -*-

"""Utils module.

This module contains utility functions termate package

"""

def merge_terms_dict(a, b):
    for key in b.keys():
        if key in a.keys() and a[key]["dc:language"] == b[key]["dc:language"]:
            for item in b[key]["dc:uri"]:
                a[key]["dc:uri"].append(item)
            for item in b[key]["frequency"]:
                a[key]["frequency"].append(item)
        else:
            a[key] = {}
            for b_key in b[key].keys():
                a[key][b_key] = b[key][b_key]


import logging
import re
from pathlib import Path
from typing import Tuple, Union, Dict

import pyshacl
from colorama import Fore, Style
from openpyxl import load_workbook as _load_workbook
from openpyxl.workbook.workbook import Workbook
from pyshacl.pytypes import GraphLike
from rdflib import Graph, URIRef, Literal, Namespace, BNode
from rdflib.namespace import RDF, RDFS, DCTERMS, PROV, XSD, DCAT

from .profiles import *

EXCEL_FILE_ENDINGS = ["xlsx"]
RDF_FILE_ENDINGS = {
    ".ttl": "ttl",
    ".rdf": "xml",
    ".xml": "xml",
    ".json-ld": "json-ld",
    ".json": "json-ld",
    ".nt": "nt",
    ".n3": "n3",
}
KNOWN_FILE_ENDINGS = [str(x) for x in RDF_FILE_ENDINGS.keys()] + EXCEL_FILE_ENDINGS
KNOWN_TEMPLATE_VERSIONS = [
    "0.2.1",
    "0.3.0",
    "0.4.0",
    "0.4.1",
    "0.4.2",
    "0.4.3",
    "0.4.4",
    "0.5.0",
    "0.6.0",
]
LATEST_TEMPLATE = KNOWN_TEMPLATE_VERSIONS[-1]


class ConversionError(Exception):
    pass


def load_workbook(file_path: Path) -> Workbook:
    if not file_path.name.lower().endswith(tuple(EXCEL_FILE_ENDINGS)):
        raise ValueError("Files for conversion to RDF must be Excel files ending .xlsx")
    return _load_workbook(filename=str(file_path), data_only=True)


def load_template(file_path: Path) -> Workbook:
    if not file_path.name.lower().endswith(tuple(EXCEL_FILE_ENDINGS)):
        raise ValueError(
            "Template files for RDF-to-Excel conversion must be Excel files ending .xlsx"
        )

    return _load_workbook(filename=str(file_path), data_only=True)


def get_template_version(wb: Workbook) -> str:
    # try 0.4.0, 0.5.0 & 0.6.0 locations
    try:
        intro_sheet = wb["Introduction"]
        if intro_sheet["E4"].value in KNOWN_TEMPLATE_VERSIONS:  # 0.5.0, 0.6.0
            return intro_sheet["E4"].value
        if intro_sheet["J11"].value in KNOWN_TEMPLATE_VERSIONS:  # 0.4.0
            return intro_sheet["J11"].value
    except Exception:
        pass

    # try 0.2.1 & 0.3.0 locations
    try:
        # older template version
        pi = wb["program info"]
        if pi["B2"].value in KNOWN_TEMPLATE_VERSIONS:
            return pi["B2"].value
    except Exception:
        pass

    # if we get here, the template version is either unknown or can't be located
    raise Exception(
        "The version of the Excel template you are using cannot be determined"
    )


def split_and_tidy_to_strings(s: str):
    # note this may not work in list of things that contain commas. Need to consider revising
    # to allow comma-seperated values where it'll split in commas but not in things enclosed in quotes.
    if s == "" or s is None:
        return []
    else:
        return [x.strip() for x in re.split("[,\n]\s?", s.strip()) if x != ""]


def split_and_tidy_to_iris(s: str, prefixes):
    return [
        expand_namespaces(ss.strip(), prefixes) for ss in split_and_tidy_to_strings(s)
    ]


def string_is_http_iri(s: str) -> Tuple[bool, str]:
    # returns (True, None) if the string (sort of) is an IRI
    # returns (False, message) otherwise
    messages = []
    if not s.startswith("http"):
        messages.append(
            f"HTTP IRIs must start with 'http' or 'https'. Your value was '{s}'"
        )
        if ":" in s:
            messages.append(
                f"It looks like your IRI might contain a prefix, {s.split(':')[0]+':'}, that could not be expanded. "
                "Check it's present in the Prefixes sheet of your workbook"
            )

    if " " in s:
        messages.append("IRIs cannot contain spaces")

    if len(messages) > 0:
        return False, " and ".join(messages)
    else:
        return True, ""


def all_strings_in_list_are_iris(l_: []) -> Tuple[bool, str]:
    messages = []
    if l_ is not None:
        for item in l_:
            r = string_is_http_iri(item)
            if not r[0]:
                messages.append(f"Item {item} failed with messages {r[1]}")

    if len(messages) > 0:
        return False, " and ".join(messages)
    else:
        return True, ""


def expand_namespaces(s: str, prefixes: Dict[str, Namespace]) -> Union[URIRef, str]:
    for pre in prefixes.keys():
        if s.startswith(pre):
            return URIRef(s.replace(pre, prefixes[pre]))
    if s.startswith("http"):
        return URIRef(s)
    else:
        return s


def bind_namespaces(g: Graph, prefixes: Dict[str, Namespace]):
    for pre, ns in prefixes.items():
        g.bind(pre.rstrip(":"), ns)


def string_from_iri(iri):
    s = str(iri.split("/")[-1].split("#")[-1])
    s = re.sub(r"(?<=[a-z])(?=[A-Z])|(?<=[A-Z])(?=[A-Z][a-z])", " ", s)
    s = s.title()
    s = s.replace("-", " ")

    return s


def id_from_iri(iri):
    id = str(iri.split("/")[-1].split("#")[-1])
    return Literal(id, datatype=XSD.token)


def make_agent(agent_value, agent_role, prefixes, iri_of_subject) -> Graph:
    ag = Graph()
    iri = expand_namespaces(agent_value, prefixes)
    creator_iri_conv = string_is_http_iri(str(iri))
    if not creator_iri_conv[0]:
        iri = BNode()
    ag.add((iri, RDF.type, PROV.Agent))
    ag.add((iri, RDFS.label, Literal(string_from_iri(agent_value))))
    if agent_role in [DCTERMS.creator, DCTERMS.publisher, DCTERMS.rightsHolder]:
        ag.add((iri_of_subject, agent_role, iri))
    else:
        qa = BNode()
        ag.add((iri_of_subject, PROV.qualifiedAttribution, qa))
        ag.add((qa, PROV.agent, iri))
        ag.add((qa, DCAT.hadRole, agent_role))

    return ag


def make_iri(s: str, prefixes: Dict[str, Namespace]):
    iri = expand_namespaces(s, prefixes)
    iri_conv = string_is_http_iri(str(iri))
    if not iri_conv[0]:
        raise ConversionError(iri_conv[1])
    return iri


def validate_with_profile(
    data_graph: Union[GraphLike, str, bytes],
    profile="vocpub",
    error_level=1,
    message_level=1,
    log_file=None,
):
    if profile not in PROFILES.keys():
        raise ValueError(
            "The profile chosen for conversion must be one of '{}'. 'vocpub' is default".format(
                "', '".join(profiles.PROFILES.keys())
            )
        )
    allow_warnings = True if error_level > 1 else False

    # validate the RDF file
    conforms, results_graph, results_text = pyshacl.validate(
        data_graph,
        shacl_graph=str(Path(__file__).parent / "validator.vocpub.ttl"),
        allow_warnings=allow_warnings,
    )

    logging_level = logging.INFO

    if message_level == 3:
        logging_level = logging.ERROR
    elif message_level == 2:
        logging_level = logging.WARNING

    if log_file:
        logging.basicConfig(
            level=logging_level, format="%(message)s", filename=log_file, force=True
        )
    else:
        logging.basicConfig(level=logging_level, format="%(message)s")

    info_list = []
    warning_list = []
    violation_list = []

    from rdflib.namespace import RDF, SH

    for report in results_graph.subjects(RDF.type, SH.ValidationReport):
        for result in results_graph.objects(report, SH.result):
            result_dict = {}
            for p, o in results_graph.predicate_objects(result):
                if p == SH.focusNode:
                    result_dict["focusNode"] = str(o)
                elif p == SH.resultMessage:
                    result_dict["resultMessage"] = str(o)
                elif p == SH.resultSeverity:
                    result_dict["resultSeverity"] = str(o)
                elif p == SH.sourceConstraintComponent:
                    result_dict["sourceConstraintComponent"] = str(o)
                elif p == SH.sourceShape:
                    result_dict["sourceShape"] = str(o)
                elif p == SH.value:
                    result_dict["value"] = str(o)
            result_message_formatted = log_msg(result_dict, log_file)
            result_message = log_msg(result_dict, "placeholder")
            if result_dict["resultSeverity"] == str(SH.Info):
                logging.info(result_message_formatted)
                info_list.append(result_message)
            elif result_dict["resultSeverity"] == str(SH.Warning):
                logging.warning(result_message_formatted)
                warning_list.append(result_message)
            elif result_dict["resultSeverity"] == str(SH.Violation):
                logging.error(result_message_formatted)
                violation_list.append(result_message)

    if error_level == 3:
        error_messages = violation_list
    elif error_level == 2:
        error_messages = warning_list + violation_list
    else:  # error_level == 1
        error_messages = info_list + warning_list + violation_list

    if len(error_messages) > 0:
        raise ConversionError(
            f"The file you supplied is not valid according to the {profile} profile."
        )


def log_msg(result: Dict, log_file: str) -> str:
    from rdflib.namespace import SH

    formatted_msg = ""
    message = f"""Validation Result in {result['sourceConstraintComponent'].split(str(SH))[1]} ({result['sourceConstraintComponent']}):
\tSeverity: sh:{result['resultSeverity'].split(str(SH))[1]}
\tSource Shape: <{result['sourceShape']}>
\tFocus Node: <{result['focusNode']}>
\tValue Node: <{result.get('value', '')}>
\tMessage: {result['resultMessage']}
"""
    if result["resultSeverity"] == str(SH.Info):
        formatted_msg = (
            f"INFO: {message}"
            if log_file
            else Fore.BLUE + "INFO: " + Style.RESET_ALL + message
        )
    elif result["resultSeverity"] == str(SH.Warning):
        formatted_msg = (
            f"WARNING: {message}"
            if log_file
            else Fore.YELLOW + "WARNING: " + Style.RESET_ALL + message
        )
    elif result["resultSeverity"] == str(SH.Violation):
        formatted_msg = (
            f"VIOLATION: {message}"
            if log_file
            else Fore.RED + "VIOLATION: " + Style.RESET_ALL + message
        )
    return formatted_msg
